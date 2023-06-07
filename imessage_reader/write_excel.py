#!/usr/bin/env python3

"""
Write Excel file containing iMessage data (user id, text, date, service, account)
Python 3.9+
Date created: October 1st, 2020
Date modified: August 28th, 2021
"""

from datetime import datetime

import openpyxl
from openpyxl.styles import Font


class ExelWriter:
    """This class manages the export to excel.
    """

    def __init__(self, imessage_data: list, file_path: str):
        """Constructor method

        :param imessage_data: list with MessageData objects
                containing user id, text, date, service and account
        :param file_path: path to the location of the Excel file
        """
        self.imessage_data = imessage_data
        self.file_path = file_path

    def write_data(self):
        """Write data to Excel
        (user id, text, date, service, account, is_from_me)
        """

        users = []
        messages = []
        dates = []
        services = []
        accounts = []
        is_from_me = []

        for data in self.imessage_data:
            users.append(data.user_id)
            messages.append(data.text)
            dates.append(data.date)
            services.append(data.service)
            accounts.append(data.account)
            is_from_me.append(data.is_from_me)

        # Call openpyxl.Workbook() to create a new blank Excel workbook
        workbook = openpyxl.Workbook()

        # Activate a sheet
        sheet = workbook.active

        # Set a title
        sheet.title = "iMessages"

        # Set headline style
        bold16font = Font(size=16, bold=True)

        sheet["A1"] = "User ID"
        sheet["A1"].font = bold16font

        sheet["B1"] = "Message"
        sheet["B1"].font = bold16font

        sheet["C1"] = "Date"
        sheet["C1"].font = bold16font

        sheet["D1"] = "Service"
        sheet["D1"].font = bold16font

        sheet["E1"] = "Destination Caller ID"
        sheet["E1"].font = bold16font

        sheet["F1"] = "Is From Me"
        sheet["F1"].font = bold16font

        # Write users to 1st column
        users_row = 2
        for user in users:
            sheet.cell(row=users_row, column=1).value = user
            users_row += 1

        # Write messages to 2nd column
        messages_row = 2
        for message in messages:
            sheet.cell(row=messages_row, column=2).value = message
            messages_row += 1

        # Write date to 3rd column
        dates_row = 2
        for date in dates:
            sheet.cell(row=dates_row, column=3).value = date
            dates_row += 1

        # Write services to 4th column
        service_row = 2
        for service in services:
            sheet.cell(row=service_row, column=4).value = service
            service_row += 1

        # Write accounts to 5th column
        account_row = 2
        for account in accounts:
            sheet.cell(row=account_row, column=5).value = account
            account_row += 1

        # Write is_from_me to 6th column
        is_from_me_row = 2
        for from_me in is_from_me:
            sheet.cell(row=is_from_me_row, column=6).value = from_me
            is_from_me_row += 1

        # Save the workbook (Excel file)
        try:
            workbook.save(
                self.file_path + f'iMessage-Data_{datetime.now().strftime("%Y-%m-%d")}.xlsx'
            )
            print()
            print(">>> Excel file successfully created! <<<")
            print("You find the Excel file in your Documents folder.")
            print()
        except IOError as e:
            print(">>> Cannot write Excel file! <<<")
            print(e)
